"""Comps Generator — reads step2_comps_data.json and produces xlsx + summary md.

All PE/PB/PS ratios are self-calculated from price and EPS/BPS data in the JSON.
No pre-computed values from third parties are used.

Bank mode: when valuation_primary == "PB" in the JSON, switches to PB-primary
output (PB-ROE regression, bank-specific columns like NPL, NIM).
"""

from __future__ import annotations

import json
import statistics
from datetime import date, datetime
from pathlib import Path

from src.analysis.financial import calc_pe

# ── Helpers ──────────────────────────────────────────────

def _load_comps_data(workspace: Path) -> dict:
    """Load and validate step2_comps_data.json from workspace."""
    path = workspace / "step2_comps_data.json"
    if not path.exists():
        raise FileNotFoundError(
            f"step2_comps_data.json not found in {workspace}. "
            "Create it first with peer financial data."
        )
    data = json.loads(path.read_text(encoding="utf-8"))
    if "peers" not in data or not data["peers"]:
        raise ValueError("step2_comps_data.json has no peers array.")
    return data


def _safe_pe(price: float | None, eps: float | None) -> float | None:
    """Calculate PE, returning None if inputs invalid or negative result."""
    if price is None or eps is None or eps <= 0:
        return None
    return round(price / eps, 1)


def _fmt_pe(pe: float | None) -> str:
    if pe is None:
        return "N/M"
    return f"{pe:.1f}x"


_fmt_pb = _fmt_pe  # Same formatting as PE (e.g. "0.6x")


def _fmt_pct(val: float | None, show_sign: bool = True) -> str:
    if val is None:
        return "N/A"
    sign = ("+" if val > 0 else "") if show_sign else ""
    return f"{sign}{val:.1f}%"


def _fmt_pct_2(val: float | None, show_sign: bool = True) -> str:
    """Format percentage with 2 decimal places (for NPL, NIM, ROE)."""
    if val is None:
        return "N/A"
    sign = ("+" if val > 0 else "") if show_sign else ""
    return f"{sign}{val:.2f}%"


def _fmt_rev(val: float | None, ccy: str) -> str:
    if val is None:
        return "N/A"
    return f"{val:.2f}B {ccy}"


def _staleness(as_of: str | None, threshold_days: int = 90) -> str:
    """Return 'fresh', 'stale', or 'very_stale' based on as_of date."""
    if not as_of:
        return "unknown"
    try:
        d = datetime.strptime(as_of, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return "unknown"
    days = (date.today() - d).days
    if days <= threshold_days:
        return "fresh"
    elif days <= 180:
        return "stale"
    return "very_stale"


def _staleness_emoji(s: str) -> str:
    return {"fresh": "🟢", "stale": "🟡", "very_stale": "🔴"}.get(s, "⚪")


def _is_bank_mode(data: dict) -> bool:
    """Check if this is a bank comps (valuation_primary == 'PB')."""
    return data.get("valuation_primary", "PE") == "PB"


# ── Core logic ───────────────────────────────────────────

def compute_all_pe(data: dict) -> list[dict]:
    """Compute PE for all peers across FY2024A/FY2025A/FY2026E/FY2027E.

    Returns a list of dicts, one per peer, with 'name', 'ticker', pe/pb values,
    and bank-specific fields (npl, nim, roe, pb_mrq) when available.
    """
    rows = []
    bank_mode = _is_bank_mode(data)
    # Determine which fiscal years to scan
    fy_keys = ["FY2024A", "FY2025A", "FY2026E", "FY2027E"]

    for p in data["peers"]:
        fin = p.get("financials", {})
        row = {
            "name": p["name"],
            "name_zh": p.get("name_zh", ""),
            "ticker": p["ticker"],
            "market": p.get("market", ""),
            "is_target": p.get("is_target", False),
            "ccy": p["ccy"],
            "price": p["price"],
            "mcap_bn_usd": p.get("mcap_bn_usd"),
            "mcap_bn_cny": p.get("mcap_bn_cny"),
            "fy_end": p.get("fy_end", ""),
        }
        for fy in fy_keys:
            fy_data = fin.get(fy, {}) or {}
            eps = fy_data.get("eps")
            pe_result = calc_pe(p["price"], eps, label=fy)
            row[f"eps_{fy}"] = eps
            row[f"pe_{fy}"] = pe_result.get("pe")
            row[f"pe_valid_{fy}"] = pe_result.get("valid", False)
            # PB from JSON (self-calculated by the analyst)
            row[f"pb_{fy}"] = fy_data.get("pb_mrq")
            # Bank-specific fields
            if bank_mode:
                row[f"roe_{fy}"] = fy_data.get("roe_pct")
                row[f"npl_{fy}"] = fy_data.get("npl_pct")
                row[f"nim_{fy}"] = fy_data.get("nim_pct")
            # Track data freshness
            as_of = fy_data.get("as_of")
            row[f"source_{fy}"] = fy_data.get("source", "")
            row[f"as_of_{fy}"] = as_of
            row[f"freshness_{fy}"] = _staleness(as_of)

        # Revenue & margins from FY2025A
        fy25 = fin.get("FY2025A", {}) or {}
        row["revenue_25a"] = fy25.get("revenue_bn")
        row["revenue_25a_ccy"] = fy25.get("revenue_ccy", p["ccy"])
        row["rev_yoy_25a"] = fy25.get("rev_yoy")
        row["gm_pct"] = fy25.get("gm_pct")
        row["nm_pct"] = fy25.get("nm_pct")

        # Revenue growth forward
        for fy in ["FY2026E", "FY2027E"]:
            fy_data = fin.get(fy, {}) or {}
            row[f"revenue_{fy}"] = fy_data.get("revenue_bn")
            row[f"rev_yoy_{fy}"] = fy_data.get("rev_yoy")

        row["notes"] = p.get("notes", "")
        rows.append(row)
    return rows


def peer_statistics(rows: list[dict], fy: str, metric: str = "pe",
                    exclude_target: bool = True) -> dict:
    """Compute median/mean for a given metric across non-target peers.

    metric: 'pe' or 'pb'
    """
    pool = [r for r in rows if (not exclude_target or not r["is_target"])]
    vals = [r[f"{metric}_{fy}"] for r in pool if r.get(f"{metric}_{fy}") is not None]
    if not vals:
        return {"median": None, "mean": None, "n": 0}
    return {
        "median": round(statistics.median(vals), 2),
        "mean": round(statistics.mean(vals), 2),
        "n": len(vals),
    }


def target_premium(rows: list[dict], fy: str, metric: str = "pe") -> dict | None:
    """Calculate target's premium/discount vs peer median for a given metric."""
    target = next((r for r in rows if r["is_target"]), None)
    if not target:
        return None
    stats = peer_statistics(rows, fy, metric=metric, exclude_target=True)
    if stats["median"] is None or target.get(f"{metric}_{fy}") is None:
        return None
    t_val = target[f"{metric}_{fy}"]
    m_val = stats["median"]
    premium_pct = round(((t_val / m_val) - 1) * 100, 1)
    return {
        f"target_{metric}": t_val,
        f"peer_median_{metric}": m_val,
        "premium_pct": premium_pct,
        "direction": "premium" if premium_pct > 0 else "discount",
    }


# ── XLSX generation ──────────────────────────────────────

def generate_xlsx(data: dict, rows: list[dict], output_path: Path) -> Path:
    """Generate step2_comps_analysis.xlsx from computed data."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as err:
        raise ImportError("openpyxl required: pip install openpyxl") from err

    bank_mode = _is_bank_mode(data)
    primary_metric = "PB" if bank_mode else "PE"

    wb = openpyxl.Workbook()

    # Styles
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    bank_fill = PatternFill("solid", fgColor="548235")  # Green for bank sections
    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_font = Font(name="Calibri", bold=True, size=10)
    data_font = Font(name="Calibri", size=10)
    note_font = Font(name="Calibri", size=9, italic=True, color="666666")
    highlight_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    ws = wb.active
    ws.title = "Comps"
    ws.sheet_properties.tabColor = "548235" if bank_mode else "2F5496"

    col_widths = {"A": 28}
    for i, _r in enumerate(rows):
        col_letter = chr(ord("B") + i)
        col_widths[col_letter] = 20
    col_widths[chr(ord("B") + len(rows))] = 14  # Median
    col_widths[chr(ord("B") + len(rows) + 1)] = 14  # Mean
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(rows) + 1)
    target_name = data["peers"][0].get("name_zh") or data["peers"][0]["name"]
    ws["A1"] = f"{'🏦 ' if bank_mode else ''}Peer Comps — {target_name} ({data['peers'][0]['ticker']}) — {date.today().strftime('%Y-%m-%d')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="548235" if bank_mode else "2F5496")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(rows) + 1)
    bm_label = data.get("benchmark_label", f"FY2026E Forward {primary_metric}")
    if bank_mode:
        ws["A2"] = f"Benchmark: {bm_label}. PB = Price / BPS (MRQ). ROE/NPL/NIM from financial statements. source: calculated."
    else:
        ws["A2"] = f"Benchmark: {bm_label}. All PE self-calculated (Price / EPS). source: calculated."
    ws["A2"].font = note_font

    row_idx = 4
    n_peers = len(rows)
    total_cols = 1 + n_peers + 2  # label + peers + median + mean

    def _col_offset(peer_idx: int) -> int:
        return peer_idx + 2  # 1-indexed: col B = peer 0

    # ── Header ──
    fill = bank_fill if bank_mode else hdr_fill
    headers = ["Metric"] + [f"{r.get('name_zh') or r['name']}\n{r['ticker']}" for r in rows] + ["Peer\nMedian", "Peer\nMean"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_idx, column=ci, value=h)
        c.font = hdr_font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border
    ws.row_dimensions[row_idx].height = 36
    row_idx += 1

    def _section(title: str):
        nonlocal row_idx
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        c = ws.cell(row=row_idx, column=1, value=title)
        c.font = sub_font
        c.fill = sub_fill
        for ci in range(1, total_cols + 1):
            ws.cell(row=row_idx, column=ci).border = thin_border
            ws.cell(row=row_idx, column=ci).fill = sub_fill
        row_idx += 1

    def _data_row(label: str, vals: list[str | None], highlight_col: int | None = None):
        nonlocal row_idx
        c = ws.cell(row=row_idx, column=1, value=label)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.border = thin_border
        for ci, v in enumerate(vals, 2):
            c = ws.cell(row=row_idx, column=ci, value=v if v else "—")
            c.font = data_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            # Highlight target column
            target_peer_idx = next((i for i, r in enumerate(rows) if r["is_target"]), None)
            if target_peer_idx is not None and ci == _col_offset(target_peer_idx):
                c.fill = highlight_fill
        row_idx += 1

    def _median_mean_str(fy: str, metric: str) -> tuple[str, str]:
        """Get formatted median/mean for a metric."""
        stats = peer_statistics(rows, fy, metric=metric, exclude_target=True)
        fmt_fn = _fmt_pb if metric == "pb" else _fmt_pe
        med = fmt_fn(stats["median"]) if stats["median"] else "N/A"
        mean = fmt_fn(stats["mean"]) if stats["mean"] else "N/A"
        return med, mean

    # ── Company Overview ──
    _section("📊 Company Overview")
    _data_row("Exchange", [r.get("market", "") for r in rows] + ["—", "—"])
    _data_row("Price (Local)", [f"{r['price']:.2f} {r['ccy']}" for r in rows] + ["—", "—"])
    if bank_mode:
        _data_row("Market Cap (¥B)", [f"¥{r.get('mcap_bn_cny', 0):.0f}B" if r.get("mcap_bn_cny") else "N/A" for r in rows] + ["—", "—"])
    else:
        _data_row("Market Cap ($B USD)", [f"${r.get('mcap_bn_usd', 0):.1f}B" if r.get("mcap_bn_usd") else "N/A" for r in rows] + ["—", "—"])
    _data_row("Fiscal Year End", [r.get("fy_end", "") for r in rows] + ["—", "—"])

    # ── Bank-specific: Asset Quality & Profitability ──
    if bank_mode:
        bm = data.get("benchmark_year", "FY2025A")
        _section("🏦 Bank Metrics — Asset Quality & Profitability")
        for fy in ["FY2024A", "FY2025A"]:
            label_suffix = " ★" if fy == bm else ""
            # NPL
            npl_vals = [_fmt_pct_2(r.get(f"npl_{fy}"), show_sign=False) for r in rows]
            npl_pool = [r[f"npl_{fy}"] for r in rows if r.get(f"npl_{fy}") is not None and not r["is_target"]]
            npl_med = _fmt_pct_2(statistics.median(npl_pool), show_sign=False) if npl_pool else "N/A"
            npl_mean = _fmt_pct_2(statistics.mean(npl_pool), show_sign=False) if npl_pool else "N/A"
            _data_row(f"NPL {fy}{label_suffix}", npl_vals + [npl_med, npl_mean])

            # NIM
            nim_vals = [_fmt_pct_2(r.get(f"nim_{fy}"), show_sign=False) for r in rows]
            nim_pool = [r[f"nim_{fy}"] for r in rows if r.get(f"nim_{fy}") is not None and not r["is_target"]]
            nim_med = _fmt_pct_2(statistics.median(nim_pool), show_sign=False) if nim_pool else "N/A"
            nim_mean = _fmt_pct_2(statistics.mean(nim_pool), show_sign=False) if nim_pool else "N/A"
            _data_row(f"NIM {fy}{label_suffix}", nim_vals + [nim_med, nim_mean])

            # ROE
            roe_vals = [_fmt_pct_2(r.get(f"roe_{fy}"), show_sign=False) for r in rows]
            roe_pool = [r[f"roe_{fy}"] for r in rows if r.get(f"roe_{fy}") is not None and not r["is_target"]]
            roe_med = _fmt_pct_2(statistics.median(roe_pool), show_sign=False) if roe_pool else "N/A"
            roe_mean = _fmt_pct_2(statistics.mean(roe_pool), show_sign=False) if roe_pool else "N/A"
            _data_row(f"ROE {fy}{label_suffix}", roe_vals + [roe_med, roe_mean])

            # Net Margin
            if fy == "FY2025A":
                nm_vals = [_fmt_pct(r.get("nm_pct"), show_sign=False) for r in rows]
                nm_pool = [r["nm_pct"] for r in rows if r.get("nm_pct") is not None and not r["is_target"]]
                nm_med = _fmt_pct(statistics.median(nm_pool), show_sign=False) if nm_pool else "N/A"
                nm_mean = _fmt_pct(statistics.mean(nm_pool), show_sign=False) if nm_pool else "N/A"
                _data_row(f"Net Margin {fy}", nm_vals + [nm_med, nm_mean])

    # ── Revenue ──
    _section("📈 Revenue (Billions, Local Currency)")
    _data_row("FY2025A Revenue", [_fmt_rev(r.get("revenue_25a"), r.get("revenue_25a_ccy", r["ccy"])) for r in rows] + ["—", "—"])
    _data_row("FY2025A YoY Growth", [_fmt_pct(r.get("rev_yoy_25a")) for r in rows] + ["—", "—"])
    _data_row("FY2026E Revenue", [_fmt_rev(r.get("revenue_FY2026E"), r.get("revenue_25a_ccy", r["ccy"])) for r in rows] + ["—", "—"])
    _data_row("FY2026E YoY Growth", [_fmt_pct(r.get("rev_yoy_FY2026E")) for r in rows] + ["—", "—"])
    _data_row("FY2027E Revenue", [_fmt_rev(r.get("revenue_FY2027E"), r.get("revenue_25a_ccy", r["ccy"])) for r in rows] + ["—", "—"])
    _data_row("FY2027E YoY Growth", [_fmt_pct(r.get("rev_yoy_FY2027E")) for r in rows] + ["—", "—"])

    # ── EPS ──
    _section("💰 EPS (Local Currency)")
    for fy in ["FY2024A", "FY2025A"]:
        _data_row(f"{fy} EPS", [_fmt_pe_val(r.get(f"eps_{fy}"), r["ccy"]) for r in rows] + ["—", "—"])
    _data_row("FY2026E EPS (Consensus)", [_fmt_pe_val(r.get("eps_FY2026E"), r["ccy"]) for r in rows] + ["—", "—"])
    _data_row("FY2027E EPS (Consensus)", [_fmt_pe_val(r.get("eps_FY2027E"), r["ccy"]) for r in rows] + ["—", "—"])

    # ── Primary Valuation Section ──
    if bank_mode:
        # PB-primary valuation
        _section("⚖️ PB Ratio (Self-Calculated: Price / BPS MRQ)")
        for fy in ["FY2024A", "FY2025A"]:
            label_suffix = " ★" if fy == data.get("benchmark_year", "FY2025A") else ""
            pb_vals = [_fmt_pb(r.get(f"pb_{fy}")) for r in rows]
            med_str, mean_str = _median_mean_str(fy, "pb")
            _data_row(f"PB {fy}{label_suffix}", pb_vals + [med_str, mean_str])

        # PE as auxiliary (still show, but secondary)
        _section("📊 PE Ratio (Auxiliary — Self-Calculated: Price ÷ EPS)")
        for fy in ["FY2024A", "FY2025A"]:
            pe_vals = [_fmt_pe(r.get(f"pe_{fy}")) for r in rows]
            med_str, mean_str = _median_mean_str(fy, "pe")
            _data_row(f"PE {fy}", pe_vals + [med_str, mean_str])
    else:
        # PE-primary valuation (original behavior)
        _section("⚖️ PE Ratio (Self-Calculated: Current Price ÷ EPS)")
        for fy in ["FY2025A", "FY2026E", "FY2027E"]:
            label_suffix = " ★" if fy == data.get("benchmark_year", "FY2026E") else ""
            pe_vals = [_fmt_pe(r.get(f"pe_{fy}")) for r in rows]
            med_str, mean_str = _median_mean_str(fy, "pe")
            _data_row(f"PE {fy}{label_suffix}", pe_vals + [med_str, mean_str])

    # ── Margins ── (skip for banks — already in bank section)
    if not bank_mode:
        _section("📊 Profitability (FY2025A)")
        _data_row("Gross Margin", [_fmt_pct(r.get("gm_pct"), show_sign=False) for r in rows] + ["—", "—"])
        _data_row("Net Margin", [_fmt_pct(r.get("nm_pct"), show_sign=False) for r in rows] + ["—", "—"])

    # ── Premium Analysis ──
    bm = data.get("benchmark_year", "FY2026E")
    metric_key = "pb" if bank_mode else "pe"
    prem = target_premium(rows, bm, metric=metric_key)
    if prem:
        _section(f"🔍 Target Premium vs Peers ({bm} {primary_metric})")
        fmt_fn = _fmt_pb if bank_mode else _fmt_pe
        _data_row(f"Target {primary_metric} ({bm})", [fmt_fn(prem[f"target_{metric_key}"])] + [""] * (n_peers + 1))
        _data_row(f"Peer Median {primary_metric} ({bm})", [fmt_fn(prem[f"peer_median_{metric_key}"])] + [""] * (n_peers + 1))
        direction = prem["direction"]
        _data_row("Target vs Peers", [f"{abs(prem['premium_pct']):.1f}% {direction} to peer median"] + [""] * (n_peers + 1))

    # ── Data Freshness ──
    _section("📅 Data Freshness")
    for fy in ["FY2026E", "FY2027E"]:
        freshness_vals = []
        for r in rows:
            f = r.get(f"freshness_{fy}", "unknown")
            emoji = _staleness_emoji(f)
            as_of = r.get(f"as_of_{fy}", "")
            if as_of:
                freshness_vals.append(f"{emoji} {as_of}")
            else:
                freshness_vals.append(f"{emoji} no data")
        _data_row(f"Data Source ({fy})", freshness_vals + ["—", "—"])

    # ── Notes ──
    row_idx += 1
    _section("📋 Notes")
    if bank_mode:
        standard_notes = [
            "1. Primary valuation: PB (Price / BPS MRQ). ROE/NPL/NIM from financial statements. Tag: source: calculated.",
            "2. Apple-to-Apple: benchmark year marked with ★. No Trailing vs Forward mixing.",
            "3. PB-ROE framework: PB correlated with sustainable ROE. Lower ROE → justified lower PB.",
            "4. Data freshness: 🟢 ≤90d, 🟡 91-180d, 🔴 >180d.",
        ]
    else:
        standard_notes = [
            "1. All PE ratios self-calculated as Current Price ÷ EPS. Tag: source: calculated.",
            "2. Apple-to-Apple: benchmark year is marked with ★. No Trailing vs Forward mixing.",
            "3. Data freshness: 🟢 ≤90d, 🟡 91-180d, 🔴 >180d. Stale data may not reflect current consensus.",
        ]
    for n in standard_notes:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
        c = ws.cell(row=row_idx, column=1, value=n)
        c.font = note_font
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # Per-peer notes
    for r in rows:
        if r.get("notes"):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
            c = ws.cell(row=row_idx, column=1, value=f"{r['ticker']}: {r['notes']}")
            c.font = note_font
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

    # ── Sheet 2: Raw Data ──
    ws2 = wb.create_sheet("Raw Data")
    ws2.sheet_properties.tabColor = "548235"
    raw_h = ["Company", "Ticker", "Metric", "FY2024A", "FY2025A", "FY2026E", "FY2027E", "Unit", "Source"]
    for ci, h in enumerate(raw_h, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = PatternFill("solid", fgColor="548235")
        c.border = thin_border
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["H"].width = 30

    rr = 2
    for r in rows:
        raw_metrics = [
            ("Revenue", {
                "FY2024A": ("revenue_25a", "revenue_25a_ccy"),  # reuse FY25 keys as fallback
                "FY2025A": ("revenue_25a", "revenue_25a_ccy"),
                "FY2026E": ("revenue_FY2026E", "revenue_25a_ccy"),
                "FY2027E": ("revenue_FY2027E", "revenue_25a_ccy"),
            }),
            ("EPS", {
                "FY2024A": ("eps_FY2024A", "ccy"),
                "FY2025A": ("eps_FY2025A", "ccy"),
                "FY2026E": ("eps_FY2026E", "ccy"),
                "FY2027E": ("eps_FY2027E", "ccy"),
            }),
            ("PE (self-calc)", {
                "FY2024A": ("pe_FY2024A", None),
                "FY2025A": ("pe_FY2025A", None),
                "FY2026E": ("pe_FY2026E", None),
                "FY2027E": ("pe_FY2027E", None),
            }),
        ]
        # Add PB and bank metrics in bank mode
        if bank_mode:
            raw_metrics.append(("PB (MRQ)", {
                "FY2024A": ("pb_FY2024A", None),
                "FY2025A": ("pb_FY2025A", None),
                "FY2026E": ("pb_FY2026E", None),
                "FY2027E": ("pb_FY2027E", None),
            }))
            raw_metrics.append(("ROE %", {
                "FY2024A": ("roe_FY2024A", None),
                "FY2025A": ("roe_FY2025A", None),
                "FY2026E": ("roe_FY2026E", None),
                "FY2027E": ("roe_FY2027E", None),
            }))
            raw_metrics.append(("NPL %", {
                "FY2024A": ("npl_FY2024A", None),
                "FY2025A": ("npl_FY2025A", None),
                "FY2026E": ("npl_FY2026E", None),
                "FY2027E": ("npl_FY2027E", None),
            }))
            raw_metrics.append(("NIM %", {
                "FY2024A": ("nim_FY2024A", None),
                "FY2025A": ("nim_FY2025A", None),
                "FY2026E": ("nim_FY2026E", None),
                "FY2027E": ("nim_FY2027E", None),
            }))

        for metric, fy_keys in raw_metrics:
            ws2.cell(row=rr, column=1, value=r["name"]).font = data_font
            ws2.cell(row=rr, column=2, value=r["ticker"]).font = data_font
            ws2.cell(row=rr, column=3, value=metric).font = data_font
            ws2.cell(row=rr, column=9, value="source: calculated" if any(k in metric for k in ("PE", "PB")) else "Financial statements").font = data_font
            for fy, (val_key, unit_key) in fy_keys.items():
                val = r.get(val_key)
                if "PE" in metric:
                    cell_val = _fmt_pe(val)
                    unit = "x"
                elif "PB" in metric:
                    cell_val = _fmt_pb(val)
                    unit = "x"
                elif "%" in metric:
                    cell_val = _fmt_pct_2(val, show_sign=False)
                    unit = "%"
                elif val is None:
                    cell_val = "N/A"
                    unit = ""
                else:
                    cell_val = f"{val:.2f}"
                    unit = r.get(unit_key, "") if unit_key else ""
                col_map = {"FY2024A": 4, "FY2025A": 5, "FY2026E": 6, "FY2027E": 7}
                ws2.cell(row=rr, column=col_map[fy], value=cell_val).font = data_font
                ws2.cell(row=rr, column=8, value=unit).font = data_font
            for ci in range(1, 10):
                ws2.cell(row=rr, column=ci).border = thin_border
            rr += 1
        rr += 1

    wb.save(str(output_path))
    return output_path


def _fmt_pe_val(eps: float | None, ccy: str) -> str:
    if eps is None:
        return "N/M (neg)"
    return f"{eps:.2f} {ccy}"


# ── Summary MD generation ────────────────────────────────

def generate_summary_md(data: dict, rows: list[dict], output_path: Path) -> Path:
    """Generate step2_comps_summary.md from computed data."""
    bank_mode = _is_bank_mode(data)
    target = next((r for r in rows if r["is_target"]), rows[0])
    bm = data.get("benchmark_year", "FY2026E")
    metric_key = "pb" if bank_mode else "pe"
    primary_label = "PB" if bank_mode else "PE"
    prem = target_premium(rows, bm, metric=metric_key)
    stats = peer_statistics(rows, bm, metric=metric_key, exclude_target=True)

    target_display = target.get("name_zh") or target["name"]

    lines = [
        f"# Step 2 Comps Summary — {target_display} ({target['ticker']})",
        "",
        f"**日期**: {date.today().strftime('%Y-%m-%d')} | **价格**: {target['price']} {target['ccy']}",
        f"**估值方法**: {primary_label}-primary{' (银行模式)' if bank_mode else ''}",
        "",
        "---",
        "",
        "## Peer Universe",
        "",
    ]

    # Peer table
    if bank_mode:
        lines.append("| 公司 | Ticker | Market | 市值(¥亿) | FY End |")
        lines.append("|:-----|:-------|:-------|:----------|:-------|")
        for r in rows:
            bold = "**" if r["is_target"] else ""
            mcap = f"¥{r.get('mcap_bn_cny', 0):.0f}亿" if r.get("mcap_bn_cny") else "N/A"
            name = r.get("name_zh") or r["name"]
            lines.append(f"| {bold}{name}{bold} | {r['ticker']} | {r.get('market', '')} | {mcap} | {r.get('fy_end', '')} |")
    else:
        lines.append("| Company | Ticker | Market | Market Cap (USD) | FY End |")
        lines.append("|:--------|:-------|:-------|:-----------------|:-------|")
        for r in rows:
            bold = "**" if r["is_target"] else ""
            mcap = f"${r.get('mcap_bn_usd', 0):.1f}B" if r.get("mcap_bn_usd") else "N/A"
            lines.append(f"| {bold}{r['name']}{bold} | {r['ticker']} | {r.get('market', '')} | {mcap} | {r.get('fy_end', '')} |")

    # Primary valuation table
    fmt_fn = _fmt_pb if bank_mode else _fmt_pe
    lines.extend(["", "---", "", f"## {bm} {primary_label} (Primary Benchmark)", ""])
    if bank_mode:
        lines.append("**Apple-to-Apple**: PB = Price / BPS (MRQ). `source: calculated`.")
    else:
        lines.append("**Apple-to-Apple**: All PE self-calculated (Current Price ÷ EPS). `source: calculated`.")
    lines.append("")
    lines.append(f"| 公司 | Price | {primary_label} |")
    lines.append("|:-----|:------|:--|")
    for r in rows:
        bold = "**" if r["is_target"] else ""
        name = r.get("name_zh") or r["name"]
        val = r.get(f"{metric_key}_{bm}")
        lines.append(f"| {bold}{name}{bold} | {r['price']:.2f} {r['ccy']} | {fmt_fn(val)} |")
    if stats["median"]:
        lines.append(f"| **Peer Median** | — | **{fmt_fn(stats['median'])}** |")
    if stats["mean"]:
        lines.append(f"| **Peer Mean** | — | **{fmt_fn(stats['mean'])}** |")

    if prem:
        lines.extend([
            "",
            f"**{target_display} {prem['direction'].title()}**: {abs(prem['premium_pct']):.1f}% vs peer median ({fmt_fn(prem[f'target_{metric_key}'])} vs {fmt_fn(prem[f'peer_median_{metric_key}'])})",
        ])

    # Bank-specific section: PB-ROE comparison
    if bank_mode:
        lines.extend(["", "---", "", "## 🏦 Bank Metrics Comparison", ""])
        lines.append("| 公司 | PB | ROE | NPL | NIM | Net Margin |")
        lines.append("|:-----|:---|:----|:----|:----|:-----------|")
        for r in rows:
            bold = "**" if r["is_target"] else ""
            name = r.get("name_zh") or r["name"]
            pb = _fmt_pb(r.get(f"pb_{bm}"))
            roe = _fmt_pct_2(r.get(f"roe_{bm}"), show_sign=False)
            npl = _fmt_pct_2(r.get(f"npl_{bm}"), show_sign=False)
            nim = _fmt_pct_2(r.get(f"nim_{bm}"), show_sign=False)
            nm = _fmt_pct(r.get("nm_pct"), show_sign=False)
            lines.append(f"| {bold}{name}{bold} | {pb} | {roe} | {npl} | {nim} | {nm} |")

        # PB-ROE regression insight
        lines.extend(["", "### PB-ROE 框架", ""])
        # Calculate simple regression: PB = a + b * ROE
        roe_pb_pairs = [(r.get(f"roe_{bm}"), r.get(f"pb_{bm}")) for r in rows
                        if r.get(f"roe_{bm}") is not None and r.get(f"pb_{bm}") is not None]
        if len(roe_pb_pairs) >= 3:
            roe_vals = [p[0] for p in roe_pb_pairs]
            pb_vals_reg = [p[1] for p in roe_pb_pairs]
            n = len(roe_vals)
            roe_mean = statistics.mean(roe_vals)
            pb_mean = statistics.mean(pb_vals_reg)
            covar = sum((roe_vals[i] - roe_mean) * (pb_vals_reg[i] - pb_mean) for i in range(n))
            roe_var = sum((v - roe_mean) ** 2 for v in roe_vals)
            if roe_var > 0:
                slope = covar / roe_var
                intercept = pb_mean - slope * roe_mean
                r_sq_num = covar ** 2
                r_sq_den = roe_var * sum((v - pb_mean) ** 2 for v in pb_vals_reg)
                r_squared = r_sq_num / r_sq_den if r_sq_den > 0 else 0
                lines.append(f"**线性回归**: PB = {intercept:.3f} + {slope:.3f} × ROE (R² = {r_squared:.2f})")
                # Implied PB for target
                target_roe = target.get(f"roe_{bm}")
                if target_roe:
                    implied_pb = intercept + slope * target_roe
                    actual_pb = target.get(f"pb_{bm}")
                    lines.append(f"- {target_display} 实际 PB: {actual_pb:.2f}x vs ROE隐含 PB: {implied_pb:.2f}x")
                    if actual_pb and implied_pb:
                        gap = ((actual_pb / implied_pb) - 1) * 100
                        lines.append(f"- 偏离度: {gap:+.1f}% ({'低估' if gap < 0 else '高估'} vs ROE-PB回归线)")

    # Revenue growth
    lines.extend(["", "---", "", "## Revenue Growth Comparison", ""])
    lines.append("| Company | FY2025A | FY2026E YoY | FY2027E YoY |")
    lines.append("|:--------|:--------|:-----------|:-----------|")
    for r in rows:
        bold = "**" if r["is_target"] else ""
        name = r.get("name_zh") or r["name"]
        rev_ccy = r.get("revenue_25a_ccy", r["ccy"])
        rev25 = _fmt_rev(r.get("revenue_25a"), rev_ccy)
        lines.append(f"| {bold}{name}{bold} | {rev25} | {_fmt_pct(r.get('rev_yoy_FY2026E'))} | {_fmt_pct(r.get('rev_yoy_FY2027E'))} |")

    # Moat constraint
    lines.extend(["", "---", "", "## Moat → Valuation Constraint", ""])
    lines.append("*(Copy from step2_competitive_moat.md)*")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append(f"*Generated by `python -m src.cli comps` ({'🏦 Bank PB mode' if bank_mode else 'PE mode'})*")
    lines.append("*Accompanying spreadsheet: step2_comps_analysis.xlsx*")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path


# ── Public entry point ───────────────────────────────────

def run_comps(workspace: Path) -> dict:
    """Load data, compute PE/PB, generate xlsx + summary md.

    Returns a summary dict with file paths and key metrics.
    """
    data = _load_comps_data(workspace)
    bank_mode = _is_bank_mode(data)
    rows = compute_all_pe(data)
    bm = data.get("benchmark_year", "FY2026E")
    metric_key = "pb" if bank_mode else "pe"
    prem = target_premium(rows, bm, metric=metric_key)
    stats = peer_statistics(rows, bm, metric=metric_key, exclude_target=True)

    xlsx_path = workspace / "step2_comps_analysis.xlsx"
    md_path = workspace / "step2_comps_summary.md"

    generate_xlsx(data, rows, xlsx_path)
    generate_summary_md(data, rows, md_path)

    result = {
        "xlsx": str(xlsx_path),
        "summary_md": str(md_path),
        "benchmark": bm,
        "mode": "bank_PB" if bank_mode else "PE",
        f"target_{metric_key}": prem[f"target_{metric_key}"] if prem else None,
        f"peer_median_{metric_key}": stats.get("median"),
        "premium_pct": prem["premium_pct"] if prem else None,
        "n_peers": len(rows),
        "data_date": data.get("date"),
    }
    return result
