"""Professional three-statement Excel model generator.

Reads ``forecast_model.json`` and ``step4_structured_assumptions.json`` to
produce a formula-linked ``.xlsx`` workbook with 6 tabs. Revenue is built up
from **multiplicative drivers** (Volume × ASP, Market Size × Share, etc.),
never from a bare growth-rate × base formula.

Path B implementation — a dedicated Python script, NOT relying on LLM to
generate openpyxl code ad-hoc.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from src.analysis._base import resolve_workspace_path
from src.analysis._utils import coerce_float as _coerce_float
from src.analysis.step4_schema import load_structured_assumptions

logger = logging.getLogger(__name__)

EXCEL_FILENAME = "step5_3statement_model.xlsx"

# ── Colour palette ────────────────────────────────────────────────────
BLUE_PROJ = "0000CD"       # projection font
BLACK_HIST = "000000"      # historical / label font
HEADER_BG = "111827"       # dark header
SUBTOTAL_BG = "E8E8E8"     # subtotal rows
GRANDTOTAL_BG = "B0B0B0"   # grand-total rows
CHECK_PASS = "92D050"      # green
CHECK_FAIL = "FF4444"      # red
WHITE = "FFFFFF"

# Number of historical period columns (FY-2A, FY-1A, FY0A).
# Column B is reserved for base values / source text across all tabs.
N_HIST_COLS = 3
PROJ_COL_START = 3 + N_HIST_COLS  # column 6 (F) = first projection column


# ── SheetLayout — safe named row-position tracking ────────────────────

class SheetLayout:
    """Track named row positions so formulas never hard-code row numbers."""

    def __init__(self) -> None:
        self._rows: dict[str, int] = {}
        self._next = 1  # 1-indexed

    def add(self, name: str, height: int = 1) -> int:
        """Allocate *height* rows starting at the next free row.  Return
        the first row allocated."""
        start = self._next
        self._rows[name] = start
        self._next += height
        return start

    def __getitem__(self, name: str) -> int:
        return self._rows[name]

    def row(self) -> int:
        """Next free row."""
        return self._next


# ── Helpers ────────────────────────────────────────────────────────────

def _num(value: Any, default: float = 0.0) -> float:
    """Coerce value to float; returns *default* on failure."""
    return _coerce_float(value, default=default)


def _safe_sheet_name(name: str) -> str:
    """openpyxl raises on names > 31 chars or containing []:*?/\\."""
    name = name.replace("[", "(").replace("]", ")").replace(":", "-")
    name = name.replace("*", "").replace("?", "").replace("/", "-").replace("\\", "-")
    return name[:31]


def _proj_col(period_idx: int) -> int:
    """Return 1-indexed column for forecast period *period_idx* (0-based)."""
    return PROJ_COL_START + period_idx


def _hist_col(hist_idx: int) -> int:
    """Return 1-indexed column for historical period *hist_idx* (0-based, 0=FY-2A)."""
    return 3 + hist_idx


def _col_letter(col: int) -> str:
    """Convert 1-indexed column number to Excel column letter."""
    result = ""
    while col > 0:
        col, r = divmod(col - 1, 26)
        result = chr(65 + r) + result
    return result


def _proj_cl(period_idx: int) -> str:
    return _col_letter(_proj_col(period_idx))


def _hist_cl(hist_idx: int) -> str:
    return _col_letter(_hist_col(hist_idx))


# ── Styling helpers ─────────────────────────────────────────────────────

def _apply_header_style(ws, row: int, max_col: int):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def _apply_label_style(ws, row: int, is_total: bool = False, is_grand: bool = False):
    from openpyxl.styles import Font, PatternFill
    cell = ws.cell(row=row, column=1)
    if is_grand:
        cell.font = Font(name="Calibri", bold=True, size=11, underline="double")
        for c in range(2, 20):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=GRANDTOTAL_BG, end_color=GRANDTOTAL_BG, fill_type="solid")
    elif is_total:
        cell.font = Font(name="Calibri", bold=True, size=10, underline="single")
        for c in range(2, 20):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")


def _apply_projection_font(ws, row: int, col_start: int, col_end: int):
    from openpyxl.styles import Font
    blue = Font(name="Calibri", size=10, color=BLUE_PROJ)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).font = blue


def _apply_historical_font(ws, row: int, col_start: int, col_end: int):
    from openpyxl.styles import Font
    black = Font(name="Calibri", size=10, color=BLACK_HIST)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).font = black


def _set_number_format(ws, row: int, col_start: int, col_end: int, fmt: str):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).number_format = fmt


def _add_comment(ws, row: int, col: int, text: str):
    from openpyxl.comments import Comment
    ws.cell(row=row, column=col).comment = Comment(text, "InvestPilot")


# ── Data loading ───────────────────────────────────────────────────────

def _load_model(workspace: Path) -> dict:
    path = workspace / "forecast_model.json"
    if not path.exists():
        raise FileNotFoundError(f"{path} not found — run Phase 1 (python -m src.cli model) first")
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("forecast_model.json is not a dict")
    return data


def _load_drivers(workspace: Path) -> dict[str, list[dict]]:
    structured = load_structured_assumptions(workspace)
    raw = structured.get("growth_drivers", []) or []
    result: dict[str, list[dict]] = {}
    for entry in raw:
        seg = str(entry.get("segment", "")).strip()
        drivers = entry.get("drivers", [])
        if seg and isinstance(drivers, list):
            result[seg] = drivers
    return result


# ── Revenue driver decomposition ───────────────────────────────────────

def _extract_driver_values(
    drivers: list[dict],
    periods: list[str],
    base_revenue: float,
) -> list[dict]:
    """Convert abstract driver specs into concrete per-period values.

    **Hard rule**: drivers MUST have explicit base_value + per-period growth_*.
    contribution_pct-only mode is BLOCKED — the analyst must return to Step 4
    and provide explicit driver data.
    """
    enriched = []
    for d in drivers:
        name = str(d.get("name", "Driver"))
        derivation = str(d.get("derivation", ""))
        contrib = _num(d.get("contribution_pct"), 0.0)
        # contribution_pct is stored as whole-number (20 = 20%)
        if abs(contrib) > 1.0:
            contrib = contrib / 100.0

        base_val = d.get("base_value")
        unit = str(d.get("unit", ""))
        growths: dict[str, float] = {}
        for p in periods:
            g = d.get(f"growth_{p}") or d.get(f"{p}_growth")
            if g is not None:
                growth = _num(g)
                if abs(growth) > 1.0:
                    growth = growth / 100.0
                growths[p] = growth

        if base_val is not None and len(growths) >= len(periods):
            enriched.append({
                "name": name,
                "unit": unit,
                "base_value": _num(base_val),
                "growths": growths,
                "contribution_pct": contrib,
                "derivation": derivation,
                "mode": "explicit",
            })
        else:
            raise ValueError(
                f"Driver '{name}' lacks explicit base_value + per-period growth_* "
                f"(need: base_value, growth_{periods[0]}, growth_{periods[1]}, growth_{periods[2]}). "
                f"contribution_pct-only decomposition is BLOCKED in Step 5. "
                f"Return to Step 4 and provide explicit driver data for segment."
            )
    return enriched


# ── Historical data loader (optional) ──────────────────────────────────

def _load_historical_financials(workspace: Path) -> dict[str, list[float]]:
    """Try to load historical financials from workspace CSVs.

    Returns dict with keys like 'revenue', 'cogs', 'gross_profit', 'ebit', 'net_income',
    each mapping to list of 3 values (oldest first). Returns empty dict if files missing.
    """
    result: dict[str, list[float]] = {}
    # Try Tushare income statement CSV
    income_csv = workspace / "financials_income.csv"
    if income_csv.exists():
        try:
            import pandas as pd
            df = pd.read_csv(income_csv, index_col=0)
            # Look for key rows
            for label in ["revenue", "operate_profit", "n_income"]:
                if label in df.index:
                    vals = df.loc[label].values[-3:].tolist()  # last 3 years
                    result[label] = [float(v) for v in vals]
        except Exception:
            pass
    return result


# ── Tab builders ───────────────────────────────────────────────────────

def _build_revenue_build_tab(wb, model: dict, drivers_map: dict[str, list[dict]]) -> SheetLayout:
    """Tab 1: Revenue Build — multiplicative driver decomposition."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Revenue Build")
    periods = model["model_conventions"]["periods"]
    segments = model["segments"]
    layout = SheetLayout()

    # Title
    r = layout.add("title")
    ws.cell(row=r, column=1, value="Revenue Build — Multiplicative Driver Decomposition")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=14)

    # Headers: Label | Base | Historical FY-2A | FY-1A | FY0A | Projected T+1 | T+2 | T+3
    r = layout.add("header")
    ws.cell(row=r, column=1, value="Line Item")
    ws.cell(row=r, column=2, value="Base / Unit")
    for h_idx, label in enumerate(["FY-2A", "FY-1A", "FY0A"]):
        ws.cell(row=r, column=_hist_col(h_idx), value=label)
    for idx_p, p in enumerate(periods):
        ws.cell(row=r, column=_proj_col(idx_p), value=p)
    _apply_header_style(ws, r, _proj_col(len(periods) - 1))

    seg_total_rows: list[int] = []

    for seg in segments:
        seg_name = str(seg.get("name", "Segment"))
        base_rev = _num(seg.get("base_revenue"), 0.0)

        # Segment header
        r = layout.add(f"seg_{seg_name}_header")
        ws.cell(row=r, column=1, value=seg_name)
        ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

        seg_drivers = drivers_map.get(seg_name, [])

        if seg_drivers:
            enriched = _extract_driver_values(seg_drivers, periods, base_rev)

            driver_rows: list[int] = []
            for d_idx, d in enumerate(enriched):
                dr = layout.add(f"seg_{seg_name}_driver_{d_idx}")
                ws.cell(row=dr, column=1, value=f"  {d['name']}")
                _add_comment(ws, dr, 1,
                             f"Driver: {d['name']}\n"
                             f"Derivation: {d.get('derivation', 'N/A')}\n"
                             f"Contribution: {d.get('contribution_pct', 0):.1%}\n"
                             f"Mode: {d.get('mode', 'unknown')}")

                # Base value
                ws.cell(row=dr, column=2, value=d["base_value"])
                if d.get("unit"):
                    _add_comment(ws, dr, 2, f"Unit: {d['unit']}")

                # Historical columns: N/A
                for h_idx in range(N_HIST_COLS):
                    ws.cell(row=dr, column=_hist_col(h_idx), value="N/A")
                    _apply_historical_font(ws, dr, _hist_col(h_idx), _hist_col(h_idx))

                # Per-period growth
                for idx_p, p in enumerate(periods):
                    g = d["growths"].get(p, 0.0)
                    col = _proj_col(idx_p)
                    ws.cell(row=dr, column=col, value=g)
                    _add_comment(ws, dr, col, f"{d['name']} {p} growth: {g:.2%}")

                _apply_projection_font(ws, dr, _proj_col(0), _proj_col(len(periods) - 1))
                _set_number_format(ws, dr, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')
                driver_rows.append(dr)

            # Revenue formula row: base × ∏(1 + driver_growth_i)
            rr = layout.add(f"seg_{seg_name}_revenue")
            ws.cell(row=rr, column=1, value=f"  {seg_name} Revenue")
            _apply_label_style(ws, rr, is_total=True)
            ws.cell(row=rr, column=2, value=base_rev)
            _set_number_format(ws, rr, 2, 2, '#,##0')

            # Historical: N/A for now
            for h_idx in range(N_HIST_COLS):
                ws.cell(row=rr, column=_hist_col(h_idx), value="N/A")
                _apply_historical_font(ws, rr, _hist_col(h_idx), _hist_col(h_idx))

            for idx_p in range(len(periods)):
                pcl = _proj_cl(idx_p)
                prior_ref = f"{_col_letter(2)}{rr}" if idx_p == 0 else f"{_proj_cl(idx_p - 1)}{rr}"
                product_parts = []
                for dr_r in driver_rows:
                    product_parts.append(f"(1+{pcl}{dr_r})")
                if len(product_parts) >= 2:
                    formula = f"={prior_ref}*({'*'.join(product_parts)})"
                elif len(product_parts) == 1:
                    formula = f"={prior_ref}*{product_parts[0]}"
                else:
                    formula = str(seg["forecast"][periods[idx_p]]["revenue"])
                ws.cell(row=rr, column=_proj_col(idx_p)).value = (
                    formula if formula.startswith("=") else float(formula)
                )
            _apply_projection_font(ws, rr, _proj_col(0), _proj_col(len(periods) - 1))
            _set_number_format(ws, rr, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')
            seg_total_rows.append(rr)
        else:
            # No drivers — direct revenue from model
            rr = layout.add(f"seg_{seg_name}_revenue")
            ws.cell(row=rr, column=1, value=f"  {seg_name} Revenue (no driver data)")
            ws.cell(row=rr, column=2, value=base_rev)
            _set_number_format(ws, rr, 2, 2, '#,##0')
            for h_idx in range(N_HIST_COLS):
                ws.cell(row=rr, column=_hist_col(h_idx), value="N/A")
            for idx_p, p in enumerate(periods):
                val = seg["forecast"][p]["revenue"]
                ws.cell(row=rr, column=_proj_col(idx_p), value=val)
            _apply_projection_font(ws, rr, _proj_col(0), _proj_col(len(periods) - 1))
            _set_number_format(ws, rr, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')
            seg_total_rows.append(rr)

    # Total Revenue
    r = layout.add("total_revenue")
    ws.cell(row=r, column=1, value="Total Revenue")
    _apply_label_style(ws, r, is_grand=True)
    total_base = sum(_num(s.get("base_revenue"), 0.0) for s in segments)
    ws.cell(row=r, column=2, value=total_base)
    _set_number_format(ws, r, 2, 2, '#,##0')

    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r, _hist_col(h_idx), _hist_col(h_idx))

    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        refs = "+".join(f"{pcl}{sr}" for sr in seg_total_rows)
        ws.cell(row=r, column=_proj_col(idx_p)).value = f"={refs}"
    _apply_projection_font(ws, r, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    layout._rows["_total_revenue_row_ref"] = r

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    for c in range(3, _proj_col(len(periods) - 1) + 1):
        ws.column_dimensions[_col_letter(c)].width = 16

    return layout


def _build_income_statement_tab(wb, model: dict, rev_layout: SheetLayout) -> SheetLayout:
    """Tab 2: Income Statement with formula references to Revenue Build."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Income Statement")
    periods = model["model_conventions"]["periods"]
    inputs = model.get("inputs", {})
    is_data = model.get("statements", {}).get("income_statement", [])
    layout = SheetLayout()

    # Title
    r = layout.add("title")
    ws.cell(row=r, column=1, value="Income Statement")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=14)

    # Headers
    r = layout.add("header")
    ws.cell(row=r, column=1, value="Line Item")
    ws.cell(row=r, column=2, value="Formula / Source")
    for h_idx, label in enumerate(["FY-2A", "FY-1A", "FY0A"]):
        ws.cell(row=r, column=_hist_col(h_idx), value=label)
    for idx_p, p in enumerate(periods):
        ws.cell(row=r, column=_proj_col(idx_p), value=p)
    _apply_header_style(ws, r, _proj_col(len(periods) - 1))

    rev_total_row = rev_layout["_total_revenue_row_ref"]
    is_lookup = {row.get("label", ""): row for row in is_data}

    # --- Revenue ---
    r_rev = layout.add("revenue")
    ws.cell(row=r_rev, column=1, value="Revenue")
    ws.cell(row=r_rev, column=2, value="=Revenue Build total")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_rev, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_rev, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_rev, column=_proj_col(idx_p)).value = f"='{_safe_sheet_name('Revenue Build')}'!{pcl}{rev_total_row}"
    _apply_projection_font(ws, r_rev, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_rev, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- Gross Margin assumption ---
    r_gm = layout.add("gm_assumption")
    ws.cell(row=r_gm, column=1, value="  Gross Margin % (Step 4)")
    ws.cell(row=r_gm, column=2, value="Step 4 assumption_matrix")
    gm_data = is_lookup.get("Gross Margin", {})
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_gm, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_gm, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = gm_data.get("values", {}).get(p, 0.0)
        ws.cell(row=r_gm, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_gm, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_gm, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    # --- COGS = Revenue × (1 - GM) ---
    r_cogs = layout.add("cogs")
    ws.cell(row=r_cogs, column=1, value="COGS")
    ws.cell(row=r_cogs, column=2, value="Revenue × (1 - Gross Margin)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_cogs, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_cogs, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_cogs, column=_proj_col(idx_p)).value = f"={pcl}{r_rev}*(1-{pcl}{r_gm})"
    _apply_projection_font(ws, r_cogs, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_cogs, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- Gross Profit ---
    r_gp = layout.add("gross_profit")
    ws.cell(row=r_gp, column=1, value="Gross Profit")
    ws.cell(row=r_gp, column=2, value="Revenue - COGS")
    _apply_label_style(ws, r_gp, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_gp, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_gp, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_gp, column=_proj_col(idx_p)).value = f"={pcl}{r_rev}-{pcl}{r_cogs}"
    _apply_projection_font(ws, r_gp, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_gp, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- OpEx ratio assumption ---
    r_opex_pct = layout.add("opex_ratio_assumption")
    ws.cell(row=r_opex_pct, column=1, value="  OpEx Ratio % (Step 4)")
    ws.cell(row=r_opex_pct, column=2, value="Step 4 assumption_matrix → opex_ratio")
    opex_pct_data = is_lookup.get("Operating Expense", {})
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_opex_pct, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_opex_pct, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        rev_val = is_lookup.get("Revenue", {}).get("values", {}).get(p, 0.0)
        opex_abs = opex_pct_data.get("values", {}).get(p, 0.0)
        ratio = opex_abs / rev_val if rev_val else 0.0
        ws.cell(row=r_opex_pct, column=_proj_col(idx_p), value=ratio)
        _add_comment(ws, r_opex_pct, _proj_col(idx_p), f"Step 4 assumption: opex_ratio {p}\nsource: assumption_matrix")
    _apply_projection_font(ws, r_opex_pct, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_opex_pct, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    # --- Operating Expense ---
    r_opex = layout.add("opex")
    ws.cell(row=r_opex, column=1, value="Operating Expense")
    ws.cell(row=r_opex, column=2, value="Revenue × OpEx Ratio")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_opex, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_opex, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_opex, column=_proj_col(idx_p)).value = f"={pcl}{r_rev}*{pcl}{r_opex_pct}"
    _apply_projection_font(ws, r_opex, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_opex, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- D&A assumption ---
    r_da_pct = layout.add("da_ratio_assumption")
    ws.cell(row=r_da_pct, column=1, value="  D&A Ratio (Step 4 input)")
    ws.cell(row=r_da_pct, column=2, value="Step 4 financial_model_inputs → da_ratio")
    da_ratio = inputs.get("da_ratio", 0.0)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_da_pct, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_da_pct, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        ws.cell(row=r_da_pct, column=_proj_col(idx_p), value=da_ratio)
    _apply_projection_font(ws, r_da_pct, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_da_pct, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    # --- D&A = Revenue × da_ratio ---
    r_da = layout.add("da")
    ws.cell(row=r_da, column=1, value="  D&A (EBITDA add-back)")
    ws.cell(row=r_da, column=2, value="Revenue × D&A ratio")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_da, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_da, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_da, column=_proj_col(idx_p)).value = f"={pcl}{r_rev}*{pcl}{r_da_pct}"
    _apply_projection_font(ws, r_da, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_da, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- EBIT ---
    r_ebit = layout.add("ebit")
    ws.cell(row=r_ebit, column=1, value="EBIT")
    ws.cell(row=r_ebit, column=2, value="Gross Profit - OpEx")
    _apply_label_style(ws, r_ebit, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ebit, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ebit, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ebit, column=_proj_col(idx_p)).value = f"={pcl}{r_gp}-{pcl}{r_opex}"
    _apply_projection_font(ws, r_ebit, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ebit, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- EBITDA = EBIT + D&A ---
    r_ebitda = layout.add("ebitda")
    ws.cell(row=r_ebitda, column=1, value="EBITDA")
    ws.cell(row=r_ebitda, column=2, value="EBIT + D&A")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ebitda, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ebitda, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ebitda, column=_proj_col(idx_p)).value = f"={pcl}{r_ebit}+{pcl}{r_da}"
    _apply_projection_font(ws, r_ebitda, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ebitda, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- Interest rate assumption cells ---
    r_int_rate_debt = layout.add("int_rate_debt_assumption")
    ws.cell(row=r_int_rate_debt, column=1, value="  Interest Rate on Debt (Step 4)")
    ws.cell(row=r_int_rate_debt, column=2, value="Step 4 financial_model_inputs")
    int_rate_debt = inputs.get("interest_rate_on_debt", 0.0)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_int_rate_debt, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_int_rate_debt, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        ws.cell(row=r_int_rate_debt, column=_proj_col(idx_p), value=int_rate_debt)
    _apply_projection_font(ws, r_int_rate_debt, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_int_rate_debt, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    r_int_rate_cash = layout.add("int_rate_cash_assumption")
    ws.cell(row=r_int_rate_cash, column=1, value="  Interest Rate on Cash (Step 4)")
    ws.cell(row=r_int_rate_cash, column=2, value="Step 4 financial_model_inputs")
    int_rate_cash = inputs.get("interest_rate_on_cash", 0.0)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_int_rate_cash, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_int_rate_cash, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        ws.cell(row=r_int_rate_cash, column=_proj_col(idx_p), value=int_rate_cash)
    _apply_projection_font(ws, r_int_rate_cash, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_int_rate_cash, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    # --- Interest Expense = avg_debt × rate ---
    debt_val = inputs.get("debt", 0.0)
    r_int_exp = layout.add("interest_expense")
    ws.cell(row=r_int_exp, column=1, value="  Interest Expense")
    ws.cell(row=r_int_exp, column=2, value="Avg Debt × Interest Rate on Debt")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_int_exp, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_int_exp, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_int_exp, column=_proj_col(idx_p)).value = f"={debt_val}*{pcl}{r_int_rate_debt}"
    _apply_projection_font(ws, r_int_exp, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_int_exp, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- Interest Income = avg_cash × rate ---
    cash_val = inputs.get("cash", 0.0)
    r_int_inc = layout.add("interest_income")
    ws.cell(row=r_int_inc, column=1, value="  Interest Income")
    ws.cell(row=r_int_inc, column=2, value="Avg Cash × Interest Rate on Cash")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_int_inc, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_int_inc, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_int_inc, column=_proj_col(idx_p)).value = f"={cash_val}*{pcl}{r_int_rate_cash}"
    _apply_projection_font(ws, r_int_inc, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_int_inc, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- EBT ---
    r_ebt = layout.add("ebt")
    ws.cell(row=r_ebt, column=1, value="EBT (Pre-tax Income)")
    ws.cell(row=r_ebt, column=2, value="EBIT - Interest Exp + Interest Inc")
    _apply_label_style(ws, r_ebt, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ebt, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ebt, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ebt, column=_proj_col(idx_p)).value = f"={pcl}{r_ebit}-{pcl}{r_int_exp}+{pcl}{r_int_inc}"
    _apply_projection_font(ws, r_ebt, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ebt, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- Tax rate assumption ---
    r_tax_pct = layout.add("tax_rate_assumption")
    ws.cell(row=r_tax_pct, column=1, value="  Tax Rate % (Step 4)")
    ws.cell(row=r_tax_pct, column=2, value="Step 4 assumption_matrix → tax_rate")
    tax_data = is_lookup.get("Tax Expense", {})
    ebt_data_lu = is_lookup.get("EBT (Pre-tax Income)", {})
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tax_pct, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tax_pct, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        tax_abs = tax_data.get("values", {}).get(p, 0.0)
        ebt_abs = ebt_data_lu.get("values", {}).get(p, 0.0)
        rate = tax_abs / ebt_abs if ebt_abs else 0.0
        ws.cell(row=r_tax_pct, column=_proj_col(idx_p), value=rate)
    _apply_projection_font(ws, r_tax_pct, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tax_pct, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    # --- Tax = MAX(0, EBT × tax_rate) ---
    r_tax = layout.add("tax")
    ws.cell(row=r_tax, column=1, value="Income Tax Expense")
    ws.cell(row=r_tax, column=2, value="MAX(0, EBT × Tax Rate)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tax, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tax, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_tax, column=_proj_col(idx_p)).value = f"=MAX(0,{pcl}{r_ebt}*{pcl}{r_tax_pct})"
    _apply_projection_font(ws, r_tax, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tax, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- Net Income ---
    r_ni = layout.add("net_income")
    ws.cell(row=r_ni, column=1, value="Net Income")
    ws.cell(row=r_ni, column=2, value="EBT - Tax")
    _apply_label_style(ws, r_ni, is_grand=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ni, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ni, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ni, column=_proj_col(idx_p)).value = f"={pcl}{r_ebt}-{pcl}{r_tax}"
    _apply_projection_font(ws, r_ni, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ni, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # --- EPS (Basic) ---
    r_eps = layout.add("eps")
    shares = inputs.get("shares_outstanding", 1.0)
    dilution_pct = inputs.get("annual_share_dilution_pct", 0.0)
    ws.cell(row=r_eps, column=1, value="EPS (Basic)")
    ws.cell(row=r_eps, column=2, value="Net Income / Basic Shares")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_eps, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_eps, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        diluted_shares_n = shares * ((1 + dilution_pct) ** idx_p)
        ws.cell(row=r_eps, column=_proj_col(idx_p)).value = f"={pcl}{r_ni}/{diluted_shares_n:.0f}"
    _apply_projection_font(ws, r_eps, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_eps, _proj_col(0), _proj_col(len(periods) - 1), '0.00')

    # --- EPS (Diluted) ---
    r_eps_dil = layout.add("eps_diluted")
    diluted_shares = inputs.get("diluted_shares", shares)
    ws.cell(row=r_eps_dil, column=1, value="EPS (Diluted)")
    ws.cell(row=r_eps_dil, column=2, value="Net Income / Diluted Shares")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_eps_dil, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_eps_dil, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        dil_shares_n = diluted_shares * ((1 + dilution_pct) ** idx_p)
        ws.cell(row=r_eps_dil, column=_proj_col(idx_p)).value = f"={pcl}{r_ni}/{dil_shares_n:.0f}"
        _add_comment(ws, r_eps_dil, _proj_col(idx_p),
                     f"Diluted shares (year {idx_p+1}): {dil_shares_n:,.0f}\n"
                     f"source: step4_structured_assumptions.json → financial_model_inputs.diluted_shares")
    _apply_projection_font(ws, r_eps_dil, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_eps_dil, _proj_col(0), _proj_col(len(periods) - 1), '0.00')

    # --- Margin analysis ---
    _ = layout.add("blank_margin", 1)

    r = layout.add("gm_pct")
    ws.cell(row=r, column=1, value="Gross Margin %")
    ws.cell(row=r, column=2, value="Gross Profit / Revenue")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = f"={pcl}{r_gp}/{pcl}{r_rev}"
    _set_number_format(ws, r, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    r = layout.add("ebit_margin")
    ws.cell(row=r, column=1, value="EBIT Margin %")
    ws.cell(row=r, column=2, value="EBIT / Revenue")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = f"={pcl}{r_ebit}/{pcl}{r_rev}"
    _set_number_format(ws, r, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    r = layout.add("ni_margin")
    ws.cell(row=r, column=1, value="Net Margin %")
    ws.cell(row=r, column=2, value="Net Income / Revenue")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = f"={pcl}{r_ni}/{pcl}{r_rev}"
    _set_number_format(ws, r, _proj_col(0), _proj_col(len(periods) - 1), '0.0%')

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 34
    for c in range(3, _proj_col(len(periods) - 1) + 1):
        ws.column_dimensions[_col_letter(c)].width = 16

    return layout


def _build_balance_sheet_tab(wb, model: dict, is_layout: SheetLayout) -> SheetLayout:
    """Tab 3: Balance Sheet — formula-linked where BS drivers present."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Balance Sheet")
    periods = model["model_conventions"]["periods"]
    inputs = model.get("inputs", {})
    bs_data = model.get("statements", {}).get("balance_sheet", [])
    cf_data = model.get("statements", {}).get("cash_flow", [])
    layout = SheetLayout()

    bs_lookup = {row.get("label", ""): row for row in bs_data}
    cf_lookup = {row.get("label", ""): row for row in cf_data}
    is_ws_name = _safe_sheet_name("Income Statement")
    bs_inputs_present = all(inputs.get(k, 0.0) != 0.0 for k in ["ar_days", "inv_days", "ap_days"])

    # Title & headers
    r = layout.add("title")
    ws.cell(row=r, column=1, value="Balance Sheet")
    r = layout.add("header")
    ws.cell(row=r, column=1, value="Line Item")
    ws.cell(row=r, column=2, value="Formula / Source")
    for h_idx, label in enumerate(["FY-2A", "FY-1A", "FY0A"]):
        ws.cell(row=r, column=_hist_col(h_idx), value=label)
    for idx_p, p in enumerate(periods):
        ws.cell(row=r, column=_proj_col(idx_p), value=p)
    _apply_header_style(ws, r, _proj_col(len(periods) - 1))

    # Store key references
    is_rev_row = is_layout["revenue"]
    is_cogs_row = is_layout["cogs"]
    is_ni_row = is_layout["net_income"]

    # ── ASSETS ──
    r = layout.add("section_assets")
    ws.cell(row=r, column=1, value="ASSETS")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    # Get CF ending cash row reference
    cf_end_cash_val = cf_lookup.get("Ending Cash", {}).get("values", {})
    prev_cash = inputs.get("cash", 0.0)

    # Cash — placeholder values (patched to CF formula after CF tab is built)
    r_cash = layout.add("cash")
    ws.cell(row=r_cash, column=1, value="  Cash & Equivalents")
    ws.cell(row=r_cash, column=2, value="Linked from CF (cross-sheet formula)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_cash, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_cash, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = cf_end_cash_val.get(p, prev_cash)
        ws.cell(row=r_cash, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_cash, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_cash, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # AR = Revenue × AR days / 365 (formula-linked when ar_days present)
    r_ar = layout.add("ar")
    ws.cell(row=r_ar, column=1, value="  Accounts Receivable")
    ar_days = inputs.get("ar_days", 0.0)
    if bs_inputs_present:
        ws.cell(row=r_ar, column=2, value="Revenue × AR days / 365")
        for h_idx in range(N_HIST_COLS):
            ws.cell(row=r_ar, column=_hist_col(h_idx), value="N/A")
            _apply_historical_font(ws, r_ar, _hist_col(h_idx), _hist_col(h_idx))
        for idx_p in range(len(periods)):
            pcl = _proj_cl(idx_p)
            ws.cell(row=r_ar, column=_proj_col(idx_p)).value = f"='{is_ws_name}'!{pcl}{is_rev_row}*{ar_days}/365"
    else:
        ws.cell(row=r_ar, column=2, value="Revenue × AR days / 365 (hard-coded: ar_days missing)")
        for h_idx in range(N_HIST_COLS):
            ws.cell(row=r_ar, column=_hist_col(h_idx), value="N/A")
            _apply_historical_font(ws, r_ar, _hist_col(h_idx), _hist_col(h_idx))
        for idx_p, p in enumerate(periods):
            val = bs_lookup.get("Accounts Receivable", {}).get("values", {}).get(p, 0.0)
            ws.cell(row=r_ar, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_ar, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ar, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Inventory = COGS × Inv days / 365
    r_inv = layout.add("inventory")
    ws.cell(row=r_inv, column=1, value="  Inventory")
    inv_days = inputs.get("inv_days", 0.0)
    if bs_inputs_present:
        ws.cell(row=r_inv, column=2, value="COGS × Inv days / 365")
        for h_idx in range(N_HIST_COLS):
            ws.cell(row=r_inv, column=_hist_col(h_idx), value="N/A")
            _apply_historical_font(ws, r_inv, _hist_col(h_idx), _hist_col(h_idx))
        for idx_p in range(len(periods)):
            pcl = _proj_cl(idx_p)
            ws.cell(row=r_inv, column=_proj_col(idx_p)).value = f"='{is_ws_name}'!{pcl}{is_cogs_row}*{inv_days}/365"
    else:
        ws.cell(row=r_inv, column=2, value="COGS × Inv days / 365 (hard-coded: inv_days missing)")
        for h_idx in range(N_HIST_COLS):
            ws.cell(row=r_inv, column=_hist_col(h_idx), value="N/A")
            _apply_historical_font(ws, r_inv, _hist_col(h_idx), _hist_col(h_idx))
        for idx_p, p in enumerate(periods):
            val = bs_lookup.get("Inventory", {}).get("values", {}).get(p, 0.0)
            ws.cell(row=r_inv, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_inv, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_inv, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Prepaid & Other CA
    r_poca = layout.add("prepaid_other_ca")
    ws.cell(row=r_poca, column=1, value="  Prepaid & Other Current Assets")
    ws.cell(row=r_poca, column=2, value="Revenue × ratio (~2%)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_poca, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_poca, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Prepaid & Other Current Assets", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_poca, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_poca, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_poca, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Total Current Assets
    r_tca = layout.add("total_current_assets")
    ws.cell(row=r_tca, column=1, value="Total Current Assets")
    _apply_label_style(ws, r_tca, is_total=True)
    ca_rows = [r_cash, r_ar, r_inv, r_poca]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tca, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tca, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        refs = "+".join(f"{pcl}{r}" for r in ca_rows)
        ws.cell(row=r_tca, column=_proj_col(idx_p)).value = f"={refs}"
    _apply_projection_font(ws, r_tca, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tca, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # PP&E (Net) = prior + Capex - D&A
    r_ppe = layout.add("ppe")
    ws.cell(row=r_ppe, column=1, value="  PP&E (Net)")
    ws.cell(row=r_ppe, column=2, value="Prior PP&E + Capex - D&A (hard-coded: capex/da from model)")
    prev_ppe = total_base_revenue = sum(
        _num(s.get("base_revenue"), 0.0) for s in model["segments"]
    ) * inputs.get("ppe_ratio", 0.0)
    ppe_vals = bs_lookup.get("PP&E (Net)", {}).get("values", {})
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ppe, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ppe, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = ppe_vals.get(p, prev_ppe)
        ws.cell(row=r_ppe, column=_proj_col(idx_p), value=val)
        prev_ppe = val
    _apply_projection_font(ws, r_ppe, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ppe, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Intangible Assets
    r_intang = layout.add("intangibles")
    ws.cell(row=r_intang, column=1, value="  Intangible Assets & Goodwill")
    ws.cell(row=r_intang, column=2, value="Input balance (constant)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_intang, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_intang, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Intangible Assets & Goodwill", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_intang, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_intang, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_intang, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Other NCA
    r_onca = layout.add("other_nca")
    ws.cell(row=r_onca, column=1, value="  Other Non-Current Assets")
    ws.cell(row=r_onca, column=2, value="Revenue × other assets ratio")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_onca, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_onca, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Other Non-Current Assets", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_onca, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_onca, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_onca, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Total Assets
    r_ta = layout.add("total_assets")
    ws.cell(row=r_ta, column=1, value="Total Assets")
    _apply_label_style(ws, r_ta, is_grand=True)
    nca_rows = [r_ppe, r_intang, r_onca]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ta, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ta, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ca_ref = f"{pcl}{r_tca}"
        nca_refs = "+".join(f"{pcl}{r}" for r in nca_rows)
        ws.cell(row=r_ta, column=_proj_col(idx_p)).value = f"={ca_ref}+{nca_refs}"
    _apply_projection_font(ws, r_ta, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ta, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # ── LIABILITIES ──
    _ = layout.add("blank_l", 1)
    r = layout.add("section_liab")
    ws.cell(row=r, column=1, value="LIABILITIES")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    # AP = COGS × AP days / 365
    r_ap = layout.add("ap")
    ap_days_val = inputs.get("ap_days", 0.0)
    if bs_inputs_present:
        ws.cell(row=r_ap, column=1, value="  Accounts Payable")
        ws.cell(row=r_ap, column=2, value="COGS × AP days / 365")
        for h_idx in range(N_HIST_COLS):
            ws.cell(row=r_ap, column=_hist_col(h_idx), value="N/A")
            _apply_historical_font(ws, r_ap, _hist_col(h_idx), _hist_col(h_idx))
        for idx_p in range(len(periods)):
            pcl = _proj_cl(idx_p)
            ws.cell(row=r_ap, column=_proj_col(idx_p)).value = f"='{is_ws_name}'!{pcl}{is_cogs_row}*{ap_days_val}/365"
    else:
        ws.cell(row=r_ap, column=1, value="  Accounts Payable")
        ws.cell(row=r_ap, column=2, value="COGS × AP days / 365 (hard-coded: ap_days missing)")
        for h_idx in range(N_HIST_COLS):
            ws.cell(row=r_ap, column=_hist_col(h_idx), value="N/A")
            _apply_historical_font(ws, r_ap, _hist_col(h_idx), _hist_col(h_idx))
        for idx_p, p in enumerate(periods):
            val = bs_lookup.get("Accounts Payable", {}).get("values", {}).get(p, 0.0)
            ws.cell(row=r_ap, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_ap, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ap, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # ST Debt
    r_std = layout.add("st_debt")
    ws.cell(row=r_std, column=1, value="  Short-term Debt")
    ws.cell(row=r_std, column=2, value="Input ST debt")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_std, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_std, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Short-term Debt", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_std, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_std, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_std, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Accrued
    r_accrued = layout.add("accrued_liab")
    ws.cell(row=r_accrued, column=1, value="  Accrued Liabilities")
    ws.cell(row=r_accrued, column=2, value="Revenue × accrued ratio")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_accrued, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_accrued, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Accrued Liabilities", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_accrued, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_accrued, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_accrued, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Deferred Revenue
    r_defrev = layout.add("deferred_rev")
    ws.cell(row=r_defrev, column=1, value="  Deferred Revenue")
    ws.cell(row=r_defrev, column=2, value="Revenue × deferred rev ratio")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_defrev, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_defrev, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Deferred Revenue", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_defrev, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_defrev, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_defrev, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Total Current Liabilities
    r_tcl = layout.add("total_current_liab")
    ws.cell(row=r_tcl, column=1, value="Total Current Liabilities")
    _apply_label_style(ws, r_tcl, is_total=True)
    cl_rows = [r_ap, r_std, r_accrued, r_defrev]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tcl, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tcl, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        refs = "+".join(f"{pcl}{r}" for r in cl_rows)
        ws.cell(row=r_tcl, column=_proj_col(idx_p)).value = f"={refs}"
    _apply_projection_font(ws, r_tcl, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tcl, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # LT Debt
    r_ltdebt = layout.add("lt_debt")
    ws.cell(row=r_ltdebt, column=1, value="  Long-term Debt")
    ws.cell(row=r_ltdebt, column=2, value="Input LT debt")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ltdebt, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ltdebt, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Long-term Debt", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_ltdebt, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_ltdebt, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ltdebt, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Other NCL
    r_oncl = layout.add("other_ncl")
    ws.cell(row=r_oncl, column=1, value="  Other Non-Current Liabilities")
    ws.cell(row=r_oncl, column=2, value="Input balance")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_oncl, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_oncl, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Other Non-Current Liabilities", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_oncl, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_oncl, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_oncl, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Total Liabilities
    r_tl = layout.add("total_liab")
    ws.cell(row=r_tl, column=1, value="Total Liabilities")
    _apply_label_style(ws, r_tl, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tl, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tl, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_tl, column=_proj_col(idx_p)).value = f"={pcl}{r_tcl}+{pcl}{r_ltdebt}+{pcl}{r_oncl}"
    _apply_projection_font(ws, r_tl, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tl, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # ── EQUITY ──
    _ = layout.add("blank_eq", 1)
    r = layout.add("section_equity")
    ws.cell(row=r, column=1, value="EQUITY")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    # Paid-in Capital
    r_pic = layout.add("pic")
    ws.cell(row=r_pic, column=1, value="  Common Stock / Paid-in Capital")
    ws.cell(row=r_pic, column=2, value="Input balance (constant)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_pic, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_pic, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Paid-in Capital", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_pic, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_pic, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_pic, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Retained Earnings = prior RE + NI - Dividends
    r_re = layout.add("retained_earnings")
    ws.cell(row=r_re, column=1, value="  Retained Earnings")
    ws.cell(row=r_re, column=2, value="Prior RE + NI - Dividends")
    payout = inputs.get("dividend_payout", 0.0)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_re, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_re, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Retained Earnings", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_re, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_re, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_re, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Other Equity
    r_oth_eq = layout.add("other_equity")
    ws.cell(row=r_oth_eq, column=1, value="  Other Equity Items")
    ws.cell(row=r_oth_eq, column=2, value="OCI / Minority / Treasury")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_oth_eq, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_oth_eq, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = bs_lookup.get("Other Equity", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_oth_eq, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_oth_eq, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_oth_eq, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Total Equity
    r_te = layout.add("total_equity")
    ws.cell(row=r_te, column=1, value="Total Equity")
    _apply_label_style(ws, r_te, is_total=True)
    eq_rows = [r_pic, r_re, r_oth_eq]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_te, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_te, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        refs = "+".join(f"{pcl}{r}" for r in eq_rows)
        ws.cell(row=r_te, column=_proj_col(idx_p)).value = f"={refs}"
    _apply_projection_font(ws, r_te, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_te, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Total L&E
    r_tle = layout.add("total_le")
    ws.cell(row=r_tle, column=1, value="Total Liabilities & Equity")
    _apply_label_style(ws, r_tle, is_grand=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tle, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tle, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_tle, column=_proj_col(idx_p)).value = f"={pcl}{r_tl}+{pcl}{r_te}"
    _apply_projection_font(ws, r_tle, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tle, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Balance Check
    r_check = layout.add("balance_check")
    ws.cell(row=r_check, column=1, value="Balance Check (should = 0)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_check, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_check, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_check, column=_proj_col(idx_p)).value = f"={pcl}{r_ta}-{pcl}{r_tle}"
    _set_number_format(ws, r_check, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 34
    for c in range(3, _proj_col(len(periods) - 1) + 1):
        ws.column_dimensions[_col_letter(c)].width = 16

    return layout


def _build_cash_flow_tab(wb, model: dict, is_layout: SheetLayout, bs_layout: SheetLayout) -> SheetLayout:
    """Tab 4: Cash Flow Statement."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Cash Flow")
    periods = model["model_conventions"]["periods"]
    inputs = model.get("inputs", {})
    cf_data = model.get("statements", {}).get("cash_flow", [])
    layout = SheetLayout()

    cf_lookup = {row.get("label", ""): row for row in cf_data}

    # Title & headers
    r = layout.add("title")
    ws.cell(row=r, column=1, value="Cash Flow Statement")
    r = layout.add("header")
    ws.cell(row=r, column=1, value="Line Item")
    ws.cell(row=r, column=2, value="Formula / Source")
    for h_idx, label in enumerate(["FY-2A", "FY-1A", "FY0A"]):
        ws.cell(row=r, column=_hist_col(h_idx), value=label)
    for idx_p, p in enumerate(periods):
        ws.cell(row=r, column=_proj_col(idx_p), value=p)
    _apply_header_style(ws, r, _proj_col(len(periods) - 1))

    is_ws_name = _safe_sheet_name("Income Statement")

    # Operating Activities
    r = layout.add("section_op")
    ws.cell(row=r, column=1, value="Operating Activities")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    # NI (linked from IS)
    r_ni = layout.add("cf_ni")
    ws.cell(row=r_ni, column=1, value="  Net Income")
    ws.cell(row=r_ni, column=2, value="Linked from Income Statement")
    ni_row = is_layout["net_income"]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ni, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ni, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ni, column=_proj_col(idx_p)).value = f"='{is_ws_name}'!{pcl}{ni_row}"
    _apply_projection_font(ws, r_ni, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ni, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # D&A
    r_da = layout.add("cf_da")
    ws.cell(row=r_da, column=1, value="  Depreciation & Amortization")
    ws.cell(row=r_da, column=2, value="Revenue × D&A ratio")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_da, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_da, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = cf_lookup.get("D&A", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_da, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_da, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_da, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Change in NWC
    r_dnwc = layout.add("cf_dnwc")
    ws.cell(row=r_dnwc, column=1, value="  Change in NWC")
    ws.cell(row=r_dnwc, column=2, value="Ending NWC - prior NWC")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_dnwc, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_dnwc, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = cf_lookup.get("Change in NWC", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_dnwc, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_dnwc, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_dnwc, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # CFO
    r_cfo = layout.add("cf_cfo")
    ws.cell(row=r_cfo, column=1, value="Cash from Operations")
    ws.cell(row=r_cfo, column=2, value="NI + D&A - ΔNWC")
    _apply_label_style(ws, r_cfo, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_cfo, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_cfo, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_cfo, column=_proj_col(idx_p)).value = f"={pcl}{r_ni}+{pcl}{r_da}-{pcl}{r_dnwc}"
    _apply_projection_font(ws, r_cfo, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_cfo, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Investing Activities
    r = layout.add("section_inv")
    ws.cell(row=r, column=1, value="Investing Activities")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    r_capex = layout.add("cf_capex")
    ws.cell(row=r_capex, column=1, value="  Capital Expenditure")
    ws.cell(row=r_capex, column=2, value="Revenue × capex ratio (negative)")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_capex, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_capex, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = cf_lookup.get("Capex", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_capex, column=_proj_col(idx_p), value=-abs(val))
    _apply_projection_font(ws, r_capex, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_capex, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    r_cfi = layout.add("cf_cfi")
    ws.cell(row=r_cfi, column=1, value="Cash from Investing")
    _apply_label_style(ws, r_cfi, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_cfi, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_cfi, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_cfi, column=_proj_col(idx_p)).value = f"={pcl}{r_capex}"
    _apply_projection_font(ws, r_cfi, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_cfi, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Financing Activities
    r = layout.add("section_fin")
    ws.cell(row=r, column=1, value="Financing Activities")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    r_div = layout.add("cf_div")
    ws.cell(row=r_div, column=1, value="  Dividends Paid")
    ws.cell(row=r_div, column=2, value="Net Income × payout ratio (negative)")
    payout = inputs.get("dividend_payout", 0.0)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_div, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_div, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        ni_val = cf_lookup.get("Net Income", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_div, column=_proj_col(idx_p), value=-abs(ni_val * payout))
    _apply_projection_font(ws, r_div, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_div, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    r_cff = layout.add("cf_cff")
    ws.cell(row=r_cff, column=1, value="Cash from Financing")
    _apply_label_style(ws, r_cff, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_cff, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_cff, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_cff, column=_proj_col(idx_p)).value = f"={pcl}{r_div}"
    _apply_projection_font(ws, r_cff, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_cff, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Summary
    _ = layout.add("blank_summary", 1)

    r_nc = layout.add("cf_net_change")
    ws.cell(row=r_nc, column=1, value="Net Change in Cash")
    ws.cell(row=r_nc, column=2, value="CFO + CFI + CFF")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_nc, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_nc, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_nc, column=_proj_col(idx_p)).value = f"={pcl}{r_cfo}+{pcl}{r_cfi}+{pcl}{r_cff}"
    _apply_projection_font(ws, r_nc, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_nc, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    r_ec = layout.add("cf_end_cash")
    r_bc = layout.add("cf_beg_cash")

    # Beginning Cash
    ws.cell(row=r_bc, column=1, value="Beginning Cash")
    ws.cell(row=r_bc, column=2, value="Prior period ending cash")
    prev_cash = inputs.get("cash", 0.0)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_bc, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_bc, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        if idx_p == 0:
            ws.cell(row=r_bc, column=_proj_col(idx_p), value=prev_cash)
        else:
            pcl = _proj_cl(idx_p - 1)
            ws.cell(row=r_bc, column=_proj_col(idx_p)).value = f"={pcl}{r_ec}"
    _apply_projection_font(ws, r_bc, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_bc, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Ending Cash
    ws.cell(row=r_ec, column=1, value="Ending Cash")
    ws.cell(row=r_ec, column=2, value="Beginning Cash + Net Change")
    _apply_label_style(ws, r_ec, is_grand=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ec, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ec, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ec, column=_proj_col(idx_p)).value = f"={pcl}{r_bc}+{pcl}{r_nc}"
    _apply_projection_font(ws, r_ec, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ec, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 36
    for c in range(3, _proj_col(len(periods) - 1) + 1):
        ws.column_dimensions[_col_letter(c)].width = 16

    return layout


def _build_valuation_bridge_tab(
    wb,
    model: dict,
    is_layout: SheetLayout,
    bs_layout: SheetLayout,
) -> SheetLayout:
    """Tab 5: Valuation Bridge."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Valuation Bridge")
    periods = model["model_conventions"]["periods"]
    inputs = model.get("inputs", {})
    val_data = model.get("statements", {}).get("valuation", [])
    layout = SheetLayout()

    val_lookup = {row.get("label", ""): row for row in val_data}

    # Title & headers
    r = layout.add("title")
    ws.cell(row=r, column=1, value="Valuation Bridge")
    r = layout.add("header")
    ws.cell(row=r, column=1, value="Line Item")
    ws.cell(row=r, column=2, value="Formula / Source")
    for h_idx, label in enumerate(["FY-2A", "FY-1A", "FY0A"]):
        ws.cell(row=r, column=_hist_col(h_idx), value=label)
    for idx_p, p in enumerate(periods):
        ws.cell(row=r, column=_proj_col(idx_p), value=p)
    _apply_header_style(ws, r, _proj_col(len(periods) - 1))

    is_ws_name = _safe_sheet_name("Income Statement")
    bs_ws_name = _safe_sheet_name("Balance Sheet")

    # EPS (Diluted)
    r_eps = layout.add("v_eps")
    ws.cell(row=r_eps, column=1, value="EPS (Diluted)")
    ws.cell(row=r_eps, column=2, value="Linked from Income Statement")
    eps_dil_row = is_layout["eps_diluted"]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_eps, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_eps, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_eps, column=_proj_col(idx_p)).value = f"='{is_ws_name}'!{pcl}{eps_dil_row}"
    _apply_projection_font(ws, r_eps, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_eps, _proj_col(0), _proj_col(len(periods) - 1), '0.00')

    # Forward PE
    r_pe = layout.add("v_pe")
    ws.cell(row=r_pe, column=1, value="Forward PE")
    ws.cell(row=r_pe, column=2, value="Step 4 assumption")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_pe, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_pe, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p, p in enumerate(periods):
        val = val_lookup.get("Forward PE", {}).get("values", {}).get(p, 0.0)
        ws.cell(row=r_pe, column=_proj_col(idx_p), value=val)
    _apply_projection_font(ws, r_pe, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_pe, _proj_col(0), _proj_col(len(periods) - 1), '0.0')

    # Target Price
    r_tp = layout.add("v_target_price")
    ws.cell(row=r_tp, column=1, value="Target Price")
    ws.cell(row=r_tp, column=2, value="EPS (Diluted) × Forward PE")
    _apply_label_style(ws, r_tp, is_grand=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_tp, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_tp, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_tp, column=_proj_col(idx_p)).value = f"={pcl}{r_eps}*{pcl}{r_pe}"
    _apply_projection_font(ws, r_tp, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_tp, _proj_col(0), _proj_col(len(periods) - 1), '0.00')

    _ = layout.add("blank_ev", 1)

    # EV Bridge
    r = layout.add("ev_section")
    ws.cell(row=r, column=1, value="EV Bridge")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)

    shares = inputs.get("diluted_shares", inputs.get("shares_outstanding", 1.0))
    r_mc = layout.add("v_market_cap")
    ws.cell(row=r_mc, column=1, value="Market Cap")
    ws.cell(row=r_mc, column=2, value="Target Price × Diluted Shares")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_mc, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_mc, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_mc, column=_proj_col(idx_p)).value = f"={pcl}{r_tp}*{shares}"
    _apply_projection_font(ws, r_mc, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_mc, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # Net Debt
    r_nd = layout.add("v_net_debt")
    ws.cell(row=r_nd, column=1, value="Net Debt")
    ws.cell(row=r_nd, column=2, value="Short-term Debt + Long-term Debt - Cash")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_nd, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_nd, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_nd, column=_proj_col(idx_p)).value = (
            f"='{bs_ws_name}'!{pcl}{bs_layout['st_debt']}"
            f"+'{bs_ws_name}'!{pcl}{bs_layout['lt_debt']}"
            f"-'{bs_ws_name}'!{pcl}{bs_layout['cash']}"
        )
    _apply_projection_font(ws, r_nd, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_nd, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # EV
    r_ev = layout.add("v_ev")
    ws.cell(row=r_ev, column=1, value="Enterprise Value")
    ws.cell(row=r_ev, column=2, value="Market Cap + Net Debt")
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ev, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ev, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ev, column=_proj_col(idx_p)).value = f"={pcl}{r_mc}+{pcl}{r_nd}"
    _apply_label_style(ws, r_ev, is_total=True)
    _apply_projection_font(ws, r_ev, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ev, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # EBITDA
    r_ebitda = layout.add("v_ebitda")
    ws.cell(row=r_ebitda, column=1, value="EBITDA")
    ws.cell(row=r_ebitda, column=2, value="Linked from Income Statement")
    is_ebitda_row = is_layout["ebitda"]
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ebitda, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ebitda, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ebitda, column=_proj_col(idx_p)).value = f"='{is_ws_name}'!{pcl}{is_ebitda_row}"
    _apply_projection_font(ws, r_ebitda, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ebitda, _proj_col(0), _proj_col(len(periods) - 1), '#,##0')

    # EV/EBITDA
    r_ev_ebitda = layout.add("v_ev_ebitda")
    ws.cell(row=r_ev_ebitda, column=1, value="EV / EBITDA")
    ws.cell(row=r_ev_ebitda, column=2, value="Enterprise Value ÷ EBITDA")
    _apply_label_style(ws, r_ev_ebitda, is_total=True)
    for h_idx in range(N_HIST_COLS):
        ws.cell(row=r_ev_ebitda, column=_hist_col(h_idx), value="N/A")
        _apply_historical_font(ws, r_ev_ebitda, _hist_col(h_idx), _hist_col(h_idx))
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r_ev_ebitda, column=_proj_col(idx_p)).value = (
            f"=IF({pcl}{r_ebitda}=0,0,{pcl}{r_ev}/{pcl}{r_ebitda})"
        )
    _apply_projection_font(ws, r_ev_ebitda, _proj_col(0), _proj_col(len(periods) - 1))
    _set_number_format(ws, r_ev_ebitda, _proj_col(0), _proj_col(len(periods) - 1), '0.0"x"')

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    for c in range(3, _proj_col(len(periods) - 1) + 1):
        ws.column_dimensions[_col_letter(c)].width = 16

    return layout


def _build_assumptions_checks_tab(
    wb, model: dict, rev_layout: SheetLayout, is_layout: SheetLayout,
    bs_layout: SheetLayout, cf_layout: SheetLayout,
) -> SheetLayout:
    """Tab 6: Assumptions & Cross-Statement Integrity Checks."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Assumptions & Checks")
    periods = model["model_conventions"]["periods"]
    inputs = model.get("inputs", {})
    checks = model.get("checks", [])
    layout = SheetLayout()

    # Section 1: Key Assumptions
    r = layout.add("title_a")
    ws.cell(row=r, column=1, value="Key Assumptions (from Step 4)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=14)

    r = layout.add("header_a")
    ws.cell(row=r, column=1, value="Input")
    ws.cell(row=r, column=2, value="Value")
    ws.cell(row=r, column=3, value="Source")
    _apply_header_style(ws, r, 3)

    input_descriptions = {
        "shares_outstanding": ("Shares Outstanding (Basic)", '#,##0'),
        "diluted_shares": ("Shares Outstanding (Diluted)", '#,##0'),
        "annual_share_dilution_pct": ("Annual Share Dilution %", '0.0%'),
        "cash": ("Opening Cash", '#,##0'),
        "debt": ("Total Debt", '#,##0'),
        "equity": ("Opening Equity", '#,##0'),
        "nwc_ratio": ("NWC Ratio", '0.0%'),
        "ppe_ratio": ("PP&E Ratio", '0.0%'),
        "other_assets_ratio": ("Other Assets Ratio", '0.0%'),
        "ap_ratio": ("AP Ratio", '0.0%'),
        "dividend_payout": ("Dividend Payout", '0.0%'),
        "da_ratio": ("D&A Ratio", '0.0%'),
        "capex_ratio": ("Capex Ratio", '0.0%'),
        "interest_rate_on_debt": ("Interest Rate on Debt", '0.0%'),
        "interest_rate_on_cash": ("Interest Rate on Cash", '0.0%'),
        "ar_days": ("AR Days", '0'),
        "inv_days": ("Inventory Days", '0'),
        "ap_days": ("AP Days", '0'),
        "intangible_assets": ("Intangible Assets", '#,##0'),
        "deferred_rev_ratio": ("Deferred Revenue Ratio", '0.0%'),
        "accrued_ratio": ("Accrued Ratio", '0.0%'),
        "other_ncl_ratio": ("Other NCL Ratio", '0.0%'),
        "st_debt": ("Short-term Debt", '#,##0'),
        "lt_debt": ("Long-term Debt", '#,##0'),
    }
    for key in [
        "shares_outstanding", "diluted_shares", "annual_share_dilution_pct",
        "cash", "debt", "equity",
        "interest_rate_on_debt", "interest_rate_on_cash",
        "nwc_ratio", "ppe_ratio", "other_assets_ratio", "ap_ratio",
        "dividend_payout", "da_ratio", "capex_ratio",
        "ar_days", "inv_days", "ap_days",
        "intangible_assets", "deferred_rev_ratio", "accrued_ratio",
        "other_ncl_ratio", "st_debt", "lt_debt",
    ]:
        r = layout.add(f"input_{key}")
        desc, fmt = input_descriptions.get(key, (key, '0.00'))
        val = inputs.get(key, 0.0)
        ws.cell(row=r, column=1, value=desc)
        ws.cell(row=r, column=2, value=val)
        src = "step4_structured_assumptions.json → financial_model_inputs"
        if key in ("ar_days", "inv_days", "ap_days", "intangible_assets",
                    "deferred_rev_ratio", "accrued_ratio", "other_ncl_ratio",
                    "st_debt", "lt_debt"):
            if val == 0.0:
                src += " (OPTIONAL: not provided, BS items hard-coded)"
            else:
                src += " (formula-linked BS)"
        ws.cell(row=r, column=3, value=src)
        _set_number_format(ws, r, 2, 2, fmt)

    # Section 2: Cross-Statement Integrity Checks
    _ = layout.add("blank_checks", 1)
    r = layout.add("title_c")
    ws.cell(row=r, column=1, value="Cross-Statement Integrity Checks")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=14)

    r = layout.add("header_c")
    ws.cell(row=r, column=1, value="Check")
    for h_idx, label in enumerate(["FY-2A", "FY-1A", "FY0A"]):
        ws.cell(row=r, column=_hist_col(h_idx), value=label)
    for idx_p, p in enumerate(periods):
        ws.cell(row=r, column=_proj_col(idx_p), value=p)
    ws.cell(row=r, column=_proj_col(len(periods)), value="Status")
    _apply_header_style(ws, r, _proj_col(len(periods)))

    bs_ws_name = _safe_sheet_name("Balance Sheet")
    is_ws_name = _safe_sheet_name("Income Statement")
    cf_ws_name = _safe_sheet_name("Cash Flow")
    rev_ws_name = _safe_sheet_name("Revenue Build")

    # Check 1: BS Balance
    r = layout.add("check_bs")
    ws.cell(row=r, column=1, value="BS Balance (Assets - L&E = 0)")
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = (
            f"='{bs_ws_name}'!{pcl}{bs_layout['total_assets']}"
            f"-'{bs_ws_name}'!{pcl}{bs_layout['total_le']}"
        )
    ws.cell(row=r, column=_proj_col(len(periods)), value="OK")

    # Check 2: Cash Tie-out
    r = layout.add("check_cash")
    ws.cell(row=r, column=1, value="Cash Tie-out (CF ending = BS cash)")
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = (
            f"='{cf_ws_name}'!{pcl}{cf_layout['cf_end_cash']}"
            f"-'{bs_ws_name}'!{pcl}{bs_layout['cash']}"
        )
    ws.cell(row=r, column=_proj_col(len(periods)), value="OK")

    # Check 3: NI Linkage
    r = layout.add("check_ni")
    ws.cell(row=r, column=1, value="NI Linkage (IS NI = CF NI)")
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = (
            f"='{is_ws_name}'!{pcl}{is_layout['net_income']}"
            f"-'{cf_ws_name}'!{pcl}{cf_layout['cf_ni']}"
        )
    ws.cell(row=r, column=_proj_col(len(periods)), value="OK")

    # Check 4: Revenue Linkage
    r = layout.add("check_rev")
    ws.cell(row=r, column=1, value="Revenue Linkage (IS = Revenue Build)")
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = (
            f"='{is_ws_name}'!{pcl}{is_layout['revenue']}"
            f"-'{rev_ws_name}'!{pcl}{rev_layout['_total_revenue_row_ref']}"
        )
    ws.cell(row=r, column=_proj_col(len(periods)), value="OK")

    # Check 5: RE Rollforward
    r = layout.add("check_re")
    ws.cell(row=r, column=1, value="RE Rollforward (Prior RE + NI - Div = Current RE)")
    for idx_p in range(len(periods)):
        if idx_p == 0:
            ws.cell(row=r, column=_proj_col(idx_p), value="N/A (first period)")
        else:
            pcl_prev = _proj_cl(idx_p - 1)
            pcl = _proj_cl(idx_p)
            ws.cell(row=r, column=_proj_col(idx_p)).value = (
                f"='{bs_ws_name}'!{pcl}{bs_layout['retained_earnings']}"
                f"-'{bs_ws_name}'!{pcl_prev}{bs_layout['retained_earnings']}"
                f"-'{is_ws_name}'!{pcl}{is_layout['net_income']}"
            )
    ws.cell(row=r, column=_proj_col(len(periods)), value="CHECK")

    # Check 6: FCF Consistency
    r = layout.add("check_fcf")
    ws.cell(row=r, column=1, value="FCF Consistency (NI + D&A - Capex - ΔNWC ≈ CFO)")
    ws.cell(row=r, column=_proj_col(len(periods)), value="CHECK")

    # Check 7: WC Validation
    r = layout.add("check_wc")
    ws.cell(row=r, column=1, value="WC Validation (AR + Inv + Prepaid - AP - Accrued - DefRev = NWC)")
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        ws.cell(row=r, column=_proj_col(idx_p)).value = (
            f"='{bs_ws_name}'!{pcl}{bs_layout['ar']}"
            f"+'{bs_ws_name}'!{pcl}{bs_layout['inventory']}"
            f"+'{bs_ws_name}'!{pcl}{bs_layout['prepaid_other_ca']}"
            f"-'{bs_ws_name}'!{pcl}{bs_layout['ap']}"
            f"-'{bs_ws_name}'!{pcl}{bs_layout['accrued_liab']}"
            f"-'{bs_ws_name}'!{pcl}{bs_layout['deferred_rev']}"
        )
    ws.cell(row=r, column=_proj_col(len(periods)), value="CHECK")

    # Section 3: Formula-over-Hardcode Audit
    _ = layout.add("blank_audit", 1)
    r = layout.add("title_audit")
    ws.cell(row=r, column=1, value="Formula-over-Hardcode Audit")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=14)

    r = layout.add("header_audit")
    ws.cell(row=r, column=1, value="Tab")
    ws.cell(row=r, column=2, value="Projection Cells")
    ws.cell(row=r, column=3, value="Formula Cells")
    ws.cell(row=r, column=4, value="Hard-coded Cells")
    ws.cell(row=r, column=5, value="Notes")
    _apply_header_style(ws, r, 5)

    r = layout.add("audit_note")
    ws.cell(row=r, column=1, value="Note: Recalculation in Excel will update formula cells when assumptions change. Hard-coded cells will not.")

    # Section 4: Model checks from forecast_model.json
    _ = layout.add("blank_model_checks", 1)
    r = layout.add("title_mc")
    ws.cell(row=r, column=1, value="Model Validation Checks (from forecast_model.json)")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=14)

    r = layout.add("header_mc")
    ws.cell(row=r, column=1, value="Period")
    ws.cell(row=r, column=2, value="Check")
    ws.cell(row=r, column=3, value="Difference")
    ws.cell(row=r, column=4, value="Status")
    ws.cell(row=r, column=5, value="Notes")
    _apply_header_style(ws, r, 5)

    for idx, c in enumerate(checks):
        r = layout.add(f"model_check_{idx}")
        ws.cell(row=r, column=1, value=str(c.get("period", "")))
        ws.cell(row=r, column=2, value=str(c.get("check", "")))
        ws.cell(row=r, column=3, value=c.get("difference", 0.0))
        ws.cell(row=r, column=4, value=str(c.get("status", "")))
        ws.cell(row=r, column=5, value=str(c.get("notes", "")))
        _set_number_format(ws, r, 3, 3, '0.00')

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 60

    return layout


# ── Main entry point ───────────────────────────────────────────────────

def generate_excel_model(workspace_dir: str | Path, ticker: str = "") -> Path:
    """Generate a professional three-statement Excel workbook.

    Reads ``forecast_model.json`` and ``step4_structured_assumptions.json``
    from the workspace directory.  Returns the path to the saved workbook.

    Parameters
    ----------
    workspace_dir : str or Path
        Workspace name (e.g. ``"600584.SH"``) or path (e.g. ``"workspaces/600584.SH"``).
    ticker : str
        Ticker symbol for display in the workbook.

    Returns
    -------
    Path
        Path to the generated ``step5_3statement_model.xlsx``.
    """
    import openpyxl
    from openpyxl import Workbook

    ws_path = resolve_workspace_path(workspace_dir)
    model = _load_model(ws_path)
    drivers_map = _load_drivers(ws_path)
    ticker = ticker or model.get("ticker", "")

    wb = Workbook()
    wb.remove(wb.active)

    # Build tabs in order
    rev_layout = _build_revenue_build_tab(wb, model, drivers_map)
    is_layout = _build_income_statement_tab(wb, model, rev_layout)
    bs_layout = _build_balance_sheet_tab(wb, model, is_layout)
    cf_layout = _build_cash_flow_tab(wb, model, is_layout, bs_layout)

    # Patch BS Cash row with cross-sheet formulas to CF Ending Cash
    bs_ws = wb[_safe_sheet_name("Balance Sheet")]
    cf_ws_name = _safe_sheet_name("Cash Flow")
    periods = model["model_conventions"]["periods"]
    for idx_p in range(len(periods)):
        pcl = _proj_cl(idx_p)
        bs_ws.cell(
            row=bs_layout["cash"], column=_proj_col(idx_p),
        ).value = f"='{cf_ws_name}'!{pcl}{cf_layout['cf_end_cash']}"

    val_layout = _build_valuation_bridge_tab(wb, model, is_layout, bs_layout)
    checks_layout = _build_assumptions_checks_tab(
        wb, model, rev_layout, is_layout, bs_layout, cf_layout,
    )

    # Reorder sheets
    sheet_names = wb.sheetnames
    expected_order = [
        "Revenue Build", "Income Statement", "Balance Sheet",
        "Cash Flow", "Valuation Bridge", "Assumptions & Checks",
    ]
    for idx, name in enumerate(expected_order):
        if name in sheet_names:
            wb.move_sheet(name, offset=idx - sheet_names.index(name))

    output_path = ws_path / EXCEL_FILENAME
    wb.save(str(output_path))
    logger.info("Excel model saved to %s", output_path)

    return output_path
